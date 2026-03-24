
from openpyxl import Workbook
import os
import pandas as pd

def main():
    # Define paths (assumed relative to working directory)
    quast_report_path = "QUAST_output/report.tsv"
    longreadsum_unfiltered_summary_path = "longreadsum_unfiltered_output/FASTQ_summary.txt"
    longreadsum_filtered_summary_path = "longreadsum_filtered_output/FASTQ_summary.txt"
    freebayes_vcf_path = "freebayes_output.vcf"
    combined_vcf_path = "combined_output.vcf"

    # Create a workbook
    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "Metadata"

    # Sheet 1: Metadata
    ws_meta.append(["Field", "Value"])
    yaml_files = [f for f in os.listdir('.') if f.endswith('.yaml')]
    if yaml_files:
        ws_meta.append(["Workflow Config", ", ".join(yaml_files)])
    else:
        ws_meta.append(["Workflow Config", "None found"])
    ws_meta.append(["Reference", "merged_reference.fasta"])
    ws_meta.append(["Date", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")])

    # Sheet 2: Read Quality Unfiltered (Longreadsum)
    ws_reads = wb.create_sheet(title="Read Quality Unfiltered")
    if os.path.exists(longreadsum_unfiltered_summary_path):
        with open(longreadsum_unfiltered_summary_path) as f:
            for line in f:
                if line.strip():
                    parts = line.strip().split("\t")
                    if len(parts) == 2:
                        ws_reads.append(parts)

    # Sheet 3: Read Quality Filtered (Longreadsum)
    ws_reads = wb.create_sheet(title="Read Quality Filtered")
    if os.path.exists(longreadsum_filtered_summary_path):
        with open(longreadsum_filtered_summary_path) as f:
            for line in f:
                if line.strip():
                    parts = line.strip().split("\t")
                    if len(parts) == 2:
                        ws_reads.append(parts)

    # Sheet 4: Assembly Quality (QUAST)
    ws_quast = wb.create_sheet(title="Assembly Quality")
    if os.path.exists(quast_report_path):
        df_quast = pd.read_csv(quast_report_path, sep="\t")
        ws_quast.append(list(df_quast.columns))
        for _, row in df_quast.iterrows():
            ws_quast.append(list(row))

    # Sheet 5: Variant Calling Info
    ws_variants = wb.create_sheet(title="Variant Calling Info")

    # Header depends on annotated vs raw
    prioritized = ["snpeff_merged_output/snpeff_output.vcf", "snpeff_assembly_output/snpeff_output.vcf", "snpeff_reads_output/snpeff_output.vcf"]
    vcf_path = next((p for p in prioritized if os.path.exists(p)), None)

    if vcf_path:
        ws_variants.append([
            "CHROM", "POS", "REF", "ALT", "FILTER",
            "Variant_Type", "Impact", "Gene", "Raw_ANN"
        ])

        with open(vcf_path) as vcf:
            for line in vcf:
                if line.startswith("#"):
                    continue
                parts = line.strip().split("\t")
                if len(parts) < 8:
                    continue
                chrom, pos, _, ref, alt, _, fltr, info = parts[:8]

                ann_entry = next((e for e in info.split(';') if e.startswith("ANN=")), None)
                if ann_entry:
                    try:
                        fields = ann_entry[4:].split('|')  # skip "ANN="
                        var_type = fields[1]
                        impact = fields[2]
                        gene = fields[3]
                    except IndexError:
                        var_type = impact = gene = ""
                    ws_variants.append([chrom, pos, ref, alt, fltr, var_type, impact, gene, ann_entry])
                else:
                    ws_variants.append([chrom, pos, ref, alt, fltr, "", "", "", ""])



    # Sheet 6: File Paths
    ws_paths = wb.create_sheet(title="Files and Outputs")
    output_files = [
        "merged_reference.fasta", "merged_reference.gb", "merged_reference.gff3",
        "freebayes_output.vcf", "combined_output.vcf", "liftoff_output/liftoff.gff3", 
        "QUAST_output/report.tsv", "longreadsum_filtered_output/FASTQ_summary.txt",
        "strainy_output/strain_variants.vcf"
    ]
    ws_paths.append(["Filename", "Exists?"])
    for f in output_files:
        ws_paths.append([f, "Yes" if os.path.exists(f) else "No"])


    # Save workbook
    output_excel_path = "pipeline_summary.xlsx"
    wb.save(output_excel_path)
    output_excel_path

if __name__ == "__main__":
    main()